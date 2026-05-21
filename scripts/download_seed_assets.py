from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_TIMEOUT_SECONDS = 30.0


@dataclass(frozen=True)
class AssetSource:
    kind: str
    sheet: str
    row: int
    url: str
    target_dir: Path


def resolve_xlsx_path(cli_value: str | None) -> Path:
    if cli_value:
        return Path(cli_value).expanduser().resolve()
    candidates = sorted([*ROOT.glob("*.xlsx"), *ROOT.glob("*.xlsm"), *(ROOT / "data").glob("*.xlsx"), *(ROOT / "data").glob("*.xlsm")])
    if candidates:
        return candidates[0]
    raise FileNotFoundError("未找到 xlsx/xlsm，请通过 --xlsx 指定数据文件")


def normalize_url(value: object) -> str:
    return str(value or "").strip()


def column_index_by_header(headers: tuple[object, ...], required_header: str) -> int:
    for index, value in enumerate(headers):
        if str(value or "").strip() == required_header:
            return index
    raise ValueError(f"未找到列：{required_header}")


def collect_sources(workbook_path: Path, output_dir: Path) -> list[AssetSource]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "手图" not in workbook.sheetnames:
        raise ValueError("未找到 sheet：手图")
    if "款式图" not in workbook.sheetnames:
        raise ValueError("未找到 sheet：款式图")

    sources: list[AssetSource] = []
    hand_sheet = workbook["手图"]
    hand_headers = next(hand_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    hand_url_index = column_index_by_header(hand_headers, "手图URL")
    for row_number, row in enumerate(hand_sheet.iter_rows(min_row=2, values_only=True), start=2):
        url = normalize_url(row[hand_url_index] if hand_url_index < len(row) else "")
        if url:
            sources.append(AssetSource("hand", "手图", row_number, url, output_dir / "hands"))

    style_sheet = workbook["款式图"]
    style_headers = next(style_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    enhanced_url_index = column_index_by_header(style_headers, "增强后款式图URL")
    for row_number, row in enumerate(style_sheet.iter_rows(min_row=2, values_only=True), start=2):
        url = normalize_url(row[enhanced_url_index] if enhanced_url_index < len(row) else "")
        if url:
            sources.append(AssetSource("style_nail", "款式图", row_number, url, output_dir / "nails"))

    workbook.close()
    return sources


def existing_hashes(output_dir: Path) -> dict[str, Path]:
    hashes: dict[str, Path] = {}
    for path in output_dir.rglob("*"):
        if not path.is_file():
            continue
        if path.name.endswith(".json"):
            continue
        try:
            hashes[hash_file(path)] = path
        except OSError:
            continue
    return hashes


def hash_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def guess_extension(url: str, content_type: str) -> str:
    extension = mimetypes.guess_extension(content_type.split(";")[0].strip()) if content_type else None
    if extension:
        return extension
    suffix = Path(urlparse(url).path).suffix.lower()
    return suffix or ".png"


def relative_to_root(path: Path) -> str:
    try:
        return path.resolve().relative_to(ROOT).as_posix()
    except ValueError:
        return path.resolve().as_posix()


def download_asset(source: AssetSource, client: httpx.Client, known_hashes: dict[str, Path]) -> dict[str, object]:
    response = client.get(source.url)
    response.raise_for_status()
    content_type = response.headers.get("content-type", "")
    if not content_type.startswith("image/"):
        raise ValueError(f"URL 响应不是图片: {source.url}")

    content = response.content
    sha256 = hashlib.sha256(content).hexdigest()
    if sha256 in known_hashes and known_hashes[sha256].exists():
        destination = known_hashes[sha256]
        deduped = True
    else:
        source.target_dir.mkdir(parents=True, exist_ok=True)
        destination = source.target_dir / f"{sha256[:20]}{guess_extension(source.url, content_type)}"
        destination.write_bytes(content)
        known_hashes[sha256] = destination
        deduped = False

    return {
        "kind": source.kind,
        "sheet": source.sheet,
        "row": source.row,
        "url": source.url,
        "path": relative_to_root(destination),
        "content_type": content_type,
        "sha256": sha256,
        "deduped": deduped,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Download demo seed hand images and nail style images from xlsx/xlsm")
    parser.add_argument("--xlsx", type=str, default=None, help="Path to 命题三美甲评测数据（对外版）.xlsx/.xlsm")
    parser.add_argument("--output-dir", type=str, default=str(ROOT / "data" / "seed"), help="Output directory, default: data/seed")
    args = parser.parse_args()

    workbook_path = resolve_xlsx_path(args.xlsx)
    output_dir = Path(args.output_dir).expanduser()
    if not output_dir.is_absolute():
        output_dir = (ROOT / output_dir).resolve()

    sources = collect_sources(workbook_path, output_dir)
    hand_count = sum(1 for source in sources if source.kind == "hand")
    style_count = sum(1 for source in sources if source.kind == "style_nail")

    known_hashes = existing_hashes(output_dir)
    downloaded = 0
    deduped = 0
    failed: list[dict[str, object]] = []
    assets: list[dict[str, object]] = []

    timeout = httpx.Timeout(
        DEFAULT_TIMEOUT_SECONDS,
        connect=DEFAULT_TIMEOUT_SECONDS,
        read=DEFAULT_TIMEOUT_SECONDS,
        write=DEFAULT_TIMEOUT_SECONDS,
    )
    with httpx.Client(timeout=timeout, follow_redirects=True, trust_env=False) as client:
        for position, source in enumerate(sources, start=1):
            print(f"[{position}/{len(sources)}] download {source.kind} row={source.row}", flush=True)
            try:
                asset = download_asset(source, client, known_hashes)
                assets.append(asset)
                downloaded += 1
                if asset["deduped"]:
                    deduped += 1
                print(f"  ok -> {asset['path']}", flush=True)
            except Exception as exc:
                print(f"  failed -> {exc}", flush=True)
                failed.append(
                    {
                        "kind": source.kind,
                        "sheet": source.sheet,
                        "row": source.row,
                        "url": source.url,
                        "error": str(exc),
                    }
                )

    summary = {
        "downloaded_at": datetime.now().isoformat(),
        "xlsx": str(workbook_path),
        "output_dir": relative_to_root(output_dir),
        "hand_urls": hand_count,
        "style_nail_urls": style_count,
        "total_urls": len(sources),
        "downloaded": downloaded,
        "deduped": deduped,
        "failed_count": len(failed),
        "failed_urls": failed,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if failed:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

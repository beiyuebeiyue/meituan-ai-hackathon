from __future__ import annotations

import hashlib
import json
import random
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import httpx
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.nail_style import NailStyle
from app.models.style_event_daily import StyleEventDaily
from app.services.job_log_service import JobLogService
from app.utils.files import guess_extension, public_url_for_path, relative_to_base


STYLE_PRESETS = [
    {"name": "法式裸粉", "tags": ["法式", "裸粉", "通勤", "显白"], "length": "短", "scene": "通勤"},
    {"name": "奶白简约", "tags": ["奶白", "通勤", "温柔"], "length": "短", "scene": "日常"},
    {"name": "猫眼约会", "tags": ["猫眼", "约会", "显白"], "length": "中", "scene": "约会"},
    {"name": "镜面节日", "tags": ["镜面", "节日", "亮片"], "length": "长", "scene": "节日"},
    {"name": "裸透清透", "tags": ["裸透", "温柔", "通勤"], "length": "短", "scene": "通勤"},
    {"name": "贴钻晚宴", "tags": ["贴钻", "节日", "法式"], "length": "长", "scene": "派对"},
]

TITLE_TEMPLATES = [
    "这款{phrase}美甲真的太好看了",
    "{phrase}这款美甲上手太绝了",
    "最近超爱这款{phrase}美甲",
    "这款{phrase}美甲谁做谁好看",
]

DESCRIPTION_TEMPLATES = [
    "推荐大家入手这款美甲，{scene}和日常都很合适。",
    "这款真的很显手白，{scene}做它特别出片。",
    "如果最近想换美甲，可以试试这款，温柔又耐看。",
    "这款上手很有氛围感，{scene}和拍照都不会出错。",
]


@dataclass
class DownloadRecord:
    source_url: str
    saved_path: Path
    content_type: str
    sha256: str
    deduped: bool


class SeedService:
    def __init__(self) -> None:
        self.settings = get_settings()
        self.job_logs = JobLogService()

    def import_seed_data(self, db: Session, xlsx_path: Path | None = None) -> dict[str, object]:
        workbook_path = xlsx_path or self.settings.seed_xlsx
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        hand_sheet = workbook["手图"]
        style_sheet = workbook["款式图"]

        cache_by_url: dict[str, DownloadRecord] = {}
        cache_by_hash: dict[str, DownloadRecord] = {}
        failed_urls: list[dict[str, str]] = []
        hand_count = 0
        style_original_count = 0
        style_enhanced_count = 0

        job = self.job_logs.start(db, "seed_import", message="开始导入种子数据", payload={"xlsx": str(workbook_path)})
        try:
            for hand_url, style_url in hand_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                if hand_url:
                    record = self._safe_download(
                        str(hand_url),
                        self.settings.seed_path / "hands",
                        cache_by_url,
                        cache_by_hash,
                        failed_urls,
                    )
                    if record:
                        hand_count += 1
                if style_url:
                    self._safe_download(
                        str(style_url),
                        self.settings.seed_path / "styles" / "enhanced",
                        cache_by_url,
                        cache_by_hash,
                        failed_urls,
                    )

            known_enhanced_urls: set[str] = set()
            for index, original_url, enhanced_url in style_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
                original_record = None
                enhanced_record = None
                if original_url:
                    original_record = self._safe_download(
                        str(original_url),
                        self.settings.seed_path / "styles" / "original",
                        cache_by_url,
                        cache_by_hash,
                        failed_urls,
                    )
                    if original_record:
                        style_original_count += 1
                if enhanced_url:
                    enhanced_record = self._safe_download(
                        str(enhanced_url),
                        self.settings.seed_path / "styles" / "enhanced",
                        cache_by_url,
                        cache_by_hash,
                        failed_urls,
                    )
                    if enhanced_record:
                        known_enhanced_urls.add(str(enhanced_url))
                        style_enhanced_count += 1

                if enhanced_record or original_record:
                    self._upsert_style_record(db, int(index or 0), original_url, enhanced_url, original_record, enhanced_record)

            manifest = {
                "imported_at": datetime.now(ZoneInfo(self.settings.ops_report_timezone)).isoformat(),
                "hands_count": hand_count,
                "style_original_count": style_original_count,
                "style_enhanced_count": style_enhanced_count,
                "deduped_style_count": db.query(NailStyle).count(),
                "failed_urls": failed_urls,
            }
            manifest_path = self.settings.seed_path / "manifests" / "latest.json"
            manifest_path.parent.mkdir(parents=True, exist_ok=True)
            manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
            self.job_logs.finish(db, job, status="succeeded", message="种子数据导入完成", payload=manifest)
            return manifest
        except Exception as exc:
            self.job_logs.finish(db, job, status="failed", message=str(exc), payload={"xlsx": str(workbook_path)})
            raise

    def enrich_style_metadata(self, db: Session) -> dict[str, object]:
        styles = list(db.scalars(select(NailStyle).order_by(NailStyle.created_at.asc(), NailStyle.id.asc())))
        updated = 0
        for index, style in enumerate(styles, start=1):
            preset = STYLE_PRESETS[(index - 1) % len(STYLE_PRESETS)]
            image_path = self.settings.base_dir / style.local_image_path
            dominant_colors = self._extract_dominant_colors(image_path)
            color_tags = self._color_tags(dominant_colors)
            tags = list(dict.fromkeys([*preset["tags"], *color_tags]))
            style.title = self._build_natural_title(index, preset, tags, color_tags)
            style.description = self._build_natural_description(index, preset, tags)
            style.tags_json = tags
            style.dominant_colors_json = dominant_colors
            style.style_metadata_json = {
                "color_tags": color_tags,
                "occasion_tags": [preset["scene"]],
                "length": preset["length"],
                "finish": preset["tags"][0],
            }
            style.is_trending = index <= 6 or style.popularity_score >= 18
            if style.popularity_score == 0:
                style.popularity_score = max(8.0, 36.0 - index)
            db.add(style)
            updated += 1
        db.commit()
        return {"updated_count": updated}

    def generate_demo_metrics(self, db: Session, days: int = 7) -> dict[str, object]:
        styles = list(db.scalars(select(NailStyle).order_by(NailStyle.created_at.asc())))
        randomizer = random.Random(42)
        base_date = datetime.now(ZoneInfo(self.settings.ops_report_timezone)).date()
        created_rows = 0
        for day_offset in range(days):
            stat_date = base_date - timedelta(days=days - day_offset - 1)
            for index, style in enumerate(styles, start=1):
                daily = db.scalar(
                    select(StyleEventDaily).where(
                        StyleEventDaily.style_id == style.id,
                        StyleEventDaily.stat_date == stat_date,
                    )
                )
                if daily is None:
                    daily = StyleEventDaily(style_id=style.id, stat_date=stat_date)
                impressions = 40 + index * 6 + day_offset * 9
                clicks = max(5, impressions // (3 + index % 4))
                favorites = max(1, clicks // 3)
                tryons = max(1, favorites + randomizer.randint(0, 3))
                publishes = 1 if index % 9 == 0 else 0
                daily.impressions = impressions
                daily.clicks = clicks
                daily.favorites = favorites
                daily.tryons = tryons
                daily.publishes = publishes
                daily.ctr = clicks / impressions if impressions else 0.0
                db.add(daily)
                style.popularity_score = max(style.popularity_score, clicks * 0.7 + favorites * 1.5 + tryons * 2.0)
                style.is_trending = style.popularity_score >= 18
                db.add(style)
                created_rows += 1
        db.commit()
        return {"generated_rows": created_rows, "days": days}

    def _upsert_style_record(
        self,
        db: Session,
        index: int,
        original_url: str | None,
        enhanced_url: str | None,
        original_record: DownloadRecord | None,
        enhanced_record: DownloadRecord | None,
    ) -> None:
        style = None
        if enhanced_url:
            style = db.scalar(select(NailStyle).where(NailStyle.enhanced_image_url == str(enhanced_url)))
        if style is None and original_url:
            style = db.scalar(select(NailStyle).where(NailStyle.original_image_url == str(original_url)))

        primary_record = enhanced_record or original_record
        if primary_record is None:
            return

        local_path = relative_to_base(primary_record.saved_path)
        image_url = public_url_for_path(primary_record.saved_path)
        preset = STYLE_PRESETS[(max(index, 1) - 1) % len(STYLE_PRESETS)]
        if style is None:
            style = NailStyle(
                title=self._build_natural_title(index or 1, preset, preset["tags"], []),
                description=self._build_natural_description(index or 1, preset, preset["tags"]),
                image_url=image_url,
                local_image_path=local_path,
                original_image_url=str(original_url) if original_url else None,
                enhanced_image_url=str(enhanced_url) if enhanced_url else None,
                source_type="seed_xlsx",
                tags_json=[],
                dominant_colors_json=[],
                style_metadata_json={},
                popularity_score=0.0,
                is_trending=False,
            )
        else:
            style.image_url = image_url
            style.local_image_path = local_path
            style.original_image_url = str(original_url) if original_url else style.original_image_url
            style.enhanced_image_url = str(enhanced_url) if enhanced_url else style.enhanced_image_url
        db.add(style)
        db.commit()

    def _download_image(
        self,
        source_url: str,
        target_dir: Path,
        cache_by_url: dict[str, DownloadRecord],
        cache_by_hash: dict[str, DownloadRecord],
    ) -> DownloadRecord:
        if source_url in cache_by_url:
            return cache_by_url[source_url]

        response = httpx.get(source_url, timeout=30, follow_redirects=True)
        response.raise_for_status()
        content_type = response.headers.get("content-type", "")
        if not content_type.startswith("image/"):
            raise ValueError(f"URL 响应不是图片: {source_url}")

        content = response.content
        sha256 = hashlib.sha256(content).hexdigest()
        if sha256 in cache_by_hash:
            record = DownloadRecord(
                source_url=source_url,
                saved_path=cache_by_hash[sha256].saved_path,
                content_type=content_type,
                sha256=sha256,
                deduped=True,
            )
            cache_by_url[source_url] = record
            return record

        target_dir.mkdir(parents=True, exist_ok=True)
        parsed = urlparse(source_url)
        extension = guess_extension(content_type, guess_extension(parsed.path, ".png"))
        destination = target_dir / f"{sha256[:20]}{extension}"
        destination.write_bytes(content)
        record = DownloadRecord(
            source_url=source_url,
            saved_path=destination,
            content_type=content_type,
            sha256=sha256,
            deduped=False,
        )
        cache_by_url[source_url] = record
        cache_by_hash[sha256] = record
        return record

    def _safe_download(
        self,
        source_url: str,
        target_dir: Path,
        cache_by_url: dict[str, DownloadRecord],
        cache_by_hash: dict[str, DownloadRecord],
        failed_urls: list[dict[str, str]],
    ) -> DownloadRecord | None:
        try:
            return self._download_image(source_url, target_dir, cache_by_url, cache_by_hash)
        except Exception as exc:
            failed_urls.append({"url": source_url, "error": str(exc)})
            return None

    @staticmethod
    def _extract_dominant_colors(image_path: Path) -> list[str]:
        image = Image.open(image_path).convert("RGB").resize((64, 64))
        colors = image.getcolors(maxcolors=4096) or []
        colors.sort(key=lambda item: item[0], reverse=True)
        dominant = []
        for _, rgb in colors[:3]:
            dominant.append("#%02x%02x%02x" % rgb)
        return dominant

    @staticmethod
    def _color_tags(dominant_colors: list[str]) -> list[str]:
        tags = []
        for color in dominant_colors:
            red = int(color[1:3], 16)
            green = int(color[3:5], 16)
            blue = int(color[5:7], 16)
            if red > 200 and green > 180 and blue > 180:
                tags.append("裸粉")
            elif red > 200 and green > 200 and blue > 200:
                tags.append("奶白")
            elif red > 180 and blue > 180:
                tags.append("冷调")
            elif red > 180 and green > 140:
                tags.append("暖调")
        return list(dict.fromkeys(tags)) or ["裸粉"]

    @staticmethod
    def _build_natural_title(index: int, preset: dict[str, object], tags: list[str], color_tags: list[str]) -> str:
        preferred_tags = color_tags + [str(tag) for tag in tags]
        keyword_pool = [tag for tag in preferred_tags if tag in {"裸粉", "奶白", "猫眼", "镜面", "法式", "裸透", "贴钻", "显白", "温柔"}]
        if not keyword_pool:
            keyword_pool = [str(preset["name"])]
        phrase = "".join(list(dict.fromkeys(keyword_pool[:2]))) or str(preset["name"])
        template = TITLE_TEMPLATES[(max(index, 1) - 1) % len(TITLE_TEMPLATES)]
        return template.format(phrase=phrase)

    @staticmethod
    def _build_natural_description(index: int, preset: dict[str, object], tags: list[str]) -> str:
        scene = str(preset["scene"])
        template = DESCRIPTION_TEMPLATES[(max(index, 1) - 1) % len(DESCRIPTION_TEMPLATES)]
        base = template.format(scene=scene)
        if "显白" in tags:
            return f"{base[:-1]}，而且特别显手白。"
        if "温柔" in tags:
            return f"{base[:-1]}，做出来会更有温柔感。"
        if "亮片" in tags or "贴钻" in tags:
            return f"{base[:-1]}，细节看起来会更精致。"
        return base
